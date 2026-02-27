from io import BytesIO

from openpyxl import load_workbook

from swo_aws_extension.swo.excel_report_builder import ExcelReportBuilder


def test_build_from_rows_creates_valid_excel():
    headers = ["Header 1", "Header 2"]
    rows = [["Val 1", "Val 2"], ["Val 3", "Val 4"]]
    builder = ExcelReportBuilder(headers, "TestSheet")

    result = builder.build_from_rows(rows)

    assert isinstance(result, bytes)
    wb = load_workbook(filename=BytesIO(result))
    assert "TestSheet" in wb.sheetnames
    ws = wb["TestSheet"]
    assert ws.cell(row=1, column=1).value == "Header 1"
    expected_data = [
        ("Header 1", "Header 2"),
        ("Val 1", "Val 2"),
        ("Val 3", "Val 4"),
    ]
    for row_idx, expected_row in enumerate(expected_data, start=1):
        assert ws.cell(row=row_idx, column=1).value == expected_row[0]
        assert ws.cell(row=row_idx, column=2).value == expected_row[1]
