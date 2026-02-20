import datetime as dt
from io import BytesIO

import pytest
from azure.storage.blob import BlobClient, BlobServiceClient
from openpyxl import load_workbook

from swo_aws_extension.aws.client import AWSClient
from swo_aws_extension.aws.errors import AWSError
from swo_aws_extension.flows.jobs.invitations_report_creator import (
    INVITATIONS_REPORT_HEADERS,
)
from swo_aws_extension.swo.notifications.teams import TeamsNotificationManager

MODULE = "swo_aws_extension.flows.jobs.invitations_report_creator"


@pytest.fixture
def external_mocks(mocker):
    mock_aws_client = mocker.MagicMock(spec=AWSClient)
    mock_aws_cls = mocker.patch(f"{MODULE}.AWSClient", autospec=True, return_value=mock_aws_client)

    mock_blob_client = mocker.MagicMock(spec=BlobClient)
    mock_blob_client.url = "https://acc.blob.core.windows.net/container/blob.xlsx"
    mock_blob_service = mocker.MagicMock(spec=BlobServiceClient)
    mock_blob_service.get_blob_client.return_value = mock_blob_client
    mock_blob_service.account_name = "acc"
    mock_blob_service.credential = mocker.MagicMock(account_key="key")
    mocker.patch(
        f"{MODULE}.BlobServiceClient.from_connection_string",
        autospec=True,
        return_value=mock_blob_service,
    )
    mocker.patch(f"{MODULE}.generate_blob_sas", autospec=True, return_value="sas-token")

    mock_teams = mocker.MagicMock(spec=TeamsNotificationManager)
    mocker.patch(f"{MODULE}.TeamsNotificationManager", autospec=True, return_value=mock_teams)

    return {
        "aws_cls": mock_aws_cls,
        "aws_client": mock_aws_client,
        "teams": mock_teams,
        "blob_client": mock_blob_client,
        "blob_service": mock_blob_service,
    }


def _make_invitation(
    invitation_id="rt-8lr3q6sn",
    status="Active",
    target_id="651706759263",
    source_id="111111111111",
    start=dt.datetime(2025, 1, 15, tzinfo=dt.UTC),
    end=dt.datetime(2025, 7, 15, tzinfo=dt.UTC),
):
    return {
        "Id": invitation_id,
        "Status": status,
        "Target": {"ManagementAccountId": target_id},
        "Source": {"ManagementAccountId": source_id},
        "StartTimestamp": start,
        "EndTimestamp": end,
    }


def _patch_authorizations_and_agreements(
    mocker, report_creator, authorizations, agreements, orders
):
    mocker.patch.object(report_creator, "_get_authorizations", return_value=authorizations)
    mocker.patch(f"{MODULE}.get_agreements_by_query", autospec=True, return_value=agreements)
    mocker.patch.object(report_creator, "_get_orders_by_query", autospec=True, return_value=orders)


def test_create_and_notify_teams_happy_path(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-8lr3q6sn",
            customer_roles_deployed="yes",
            channel_handshake_approved="yes",
        ),
        ordering_parameters=order_parameters_factory(
            mpa_id="111111111111",
            support_type="Partner Led Support",
        ),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-8lr3q6sn",
            customer_roles_deployed="yes",
            channel_handshake_approved="yes",
        ),
    )
    invitation = _make_invitation()
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = [invitation]

    report_creator.create_and_notify_teams()  # act

    external_mocks["blob_client"].upload_blob.assert_called_once()
    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.title == "Invitations"
    headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
    assert headers == INVITATIONS_REPORT_HEADERS
    for col in range(1, 14):
        assert ws.cell(row=1, column=col).font.bold is True
    assert ws.auto_filter.ref is not None
    assert ws.cell(row=2, column=1).value == "651706759263"
    assert ws.cell(row=2, column=2).value == "111111111111"
    assert ws.cell(row=2, column=3).value == "rt-8lr3q6sn"
    assert ws.cell(row=2, column=4).value == "Active"
    assert ws.cell(row=2, column=5).value == agreement["id"]
    assert ws.cell(row=2, column=6).value == agreement["name"]
    assert ws.cell(row=2, column=7).value == "PRD-1111-1111"
    assert ws.cell(row=2, column=9).value == "Partner Led Support"
    assert ws.cell(row=2, column=10).value == "Yes"
    assert ws.cell(row=2, column=11).value == "Yes"
    assert ws.cell(row=2, column=12).value == "2025-01-15"
    assert ws.cell(row=2, column=13).value == "2025-07-15"
    external_mocks["teams"].send_success.assert_called_once()
    call_args = external_mocks["teams"].send_success.call_args
    assert "AWS Billing Transfer Invitations Report" in call_args.args[0]
    button = call_args.kwargs["button"]
    assert button.label == "Download report"
    assert "sas-token" in button.url


def test_no_authorizations(mocker, report_creator, external_mocks):
    _patch_authorizations_and_agreements(mocker, report_creator, [], [], [])

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=1).value is None
    assert "0" in external_mocks["teams"].send_success.call_args.args[1]


def test_aws_error_sends_teams_error_and_continues(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-ok",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-ok",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
    )
    authorizations = [
        {"id": "AUT-FAIL", "externalIds": {"operations": "111"}},
        {"id": "AUT-OK", "externalIds": {"operations": "222"}},
    ]
    mocker.patch.object(report_creator, "_get_authorizations", return_value=authorizations)
    ok_client = mocker.MagicMock(spec=AWSClient)
    ok_client.get_inbound_responsibility_transfers.return_value = []
    mocker.patch(f"{MODULE}.AWSClient", side_effect=[AWSError("boom"), ok_client])
    mocker.patch(f"{MODULE}.get_agreements_by_query", return_value=[agreement])
    mocker.patch.object(report_creator, "_get_orders_by_query", return_value=[order])

    report_creator.create_and_notify_teams()  # act

    external_mocks["teams"].send_error.assert_called_once()
    external_mocks["teams"].send_success.assert_called_once()


def test_invitation_without_matching_agreement(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-other",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-other",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
    )
    invitation = _make_invitation(invitation_id="rt-unmatched")
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = [invitation]

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "651706759263"
    assert ws.cell(row=2, column=3).value == "rt-unmatched"
    assert ws.cell(row=2, column=5).value is None
    assert ws.cell(row=2, column=6).value is None


def test_agreement_without_matching_invitation(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-orphan",
            customer_roles_deployed="yes",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(mpa_id="111111111111"),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-orphan",
            customer_roles_deployed="yes",
            channel_handshake_approved="no",
        ),
    )
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = []

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=1).value is None
    assert ws.cell(row=2, column=2).value is None
    assert ws.cell(row=2, column=3).value is None
    assert ws.cell(row=2, column=4).value is None
    assert ws.cell(row=2, column=5).value is None
    assert ws.cell(row=2, column=10).value is None
    assert ws.cell(row=2, column=11).value is None


def test_invitation_with_no_dates(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-8lr3q6sn",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-8lr3q6sn",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
    )
    invitation = _make_invitation(start=None, end=None)
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = [invitation]

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=12).value is None
    assert ws.cell(row=2, column=13).value is None


def test_agreement_with_none_authorization(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-x",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(mpa_id="mpa-123", support_type="PLS"),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-x",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
    )
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = []

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=1).value is None
    assert ws.cell(row=2, column=2).value is None


def test_agreement_with_none_mpa_account_id(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-y",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-y",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
    )
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = []

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=2).value is None


def test_unmatched_invitation_with_no_dates(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-other",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(),
    )
    order = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-other",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
    )
    invitation = _make_invitation(invitation_id="rt-no-match", start=None, end=None)
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker, report_creator, authorizations, [agreement], [order]
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = [invitation]

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=12).value is None
    assert ws.cell(row=2, column=13).value is None


def test_invitation_uses_order_data_when_agreement_has_no_matching_transfer_id(
    mocker,
    report_creator,
    external_mocks,
    agreement_factory,
    order_factory,
    fulfillment_parameters_factory,
    order_parameters_factory,
):
    agreement_without_transfer_id = agreement_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="",
            customer_roles_deployed="no",
            channel_handshake_approved="no",
        ),
        ordering_parameters=order_parameters_factory(support_type="Basic Support"),
    )
    order_with_transfer_id = order_factory(
        fulfillment_parameters=fulfillment_parameters_factory(
            responsibility_transfer_id="rt-from-order",
            customer_roles_deployed="yes",
            channel_handshake_approved="yes",
        ),
        order_parameters=order_parameters_factory(support_type="Enterprise On-Ramp"),
    )
    invitation = _make_invitation(invitation_id="rt-from-order")
    authorizations = [{"id": "AUT-1", "externalIds": {"operations": "651706759263"}}]
    _patch_authorizations_and_agreements(
        mocker,
        report_creator,
        authorizations,
        [agreement_without_transfer_id],
        [order_with_transfer_id],
    )
    external_mocks["aws_client"].get_inbound_responsibility_transfers.return_value = [invitation]

    report_creator.create_and_notify_teams()  # act

    excel_bytes = external_mocks["blob_client"].upload_blob.call_args.args[0]
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=3).value == "rt-from-order"
    assert ws.cell(row=2, column=5).value == order_with_transfer_id["agreement"]["id"]
    assert ws.cell(row=2, column=6).value == order_with_transfer_id["agreement"]["name"]
    assert ws.cell(row=2, column=7).value == order_with_transfer_id["product"]["id"]
    assert ws.cell(row=2, column=9).value == "Enterprise On-Ramp"
    assert ws.cell(row=2, column=10).value == "Yes"
    assert ws.cell(row=2, column=11).value == "Yes"
