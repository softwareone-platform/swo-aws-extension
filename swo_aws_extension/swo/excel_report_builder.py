from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

type CellValue = str | float | Percent
type SheetDefinition = tuple[str, list[str], list[list[CellValue]]]

_HEADER_FONT = Font(bold=True)
_PERCENT_NUMBER_FORMAT = "0.###################%"


@dataclass(frozen=True, slots=True)
class Percent:
    """A ratio (e.g. 0.1234) rendered with Excel's native percent number format."""

    ratio: float


def _resolve_cell_value(entry: CellValue) -> str | float:
    return entry.ratio if isinstance(entry, Percent) else entry


class ExcelReportBuilder:
    """Builds an Excel workbook for the invitations report."""

    def __init__(self, headers: list[str], sheet_name: str = "Invitations"):
        self.headers = headers
        self.sheet_name = sheet_name

    def build_from_rows(self, rows: list[list[CellValue]]) -> bytes:
        """Build an Excel workbook in memory and return its raw bytes."""
        wb = Workbook()
        ws = wb.active
        if ws.title == "Sheet":
            ws.title = self.sheet_name
        self._populate_sheet(ws, self.headers, rows)
        return self._save_to_bytes(wb)

    def build_multi_sheet(self, sheets: list[SheetDefinition]) -> bytes:
        """Build an Excel workbook with multiple named sheets and return its raw bytes."""
        wb = Workbook()
        for idx, sheet_def in enumerate(sheets):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = sheet_def[0]
            self._populate_sheet(ws, sheet_def[1], sheet_def[2])
        return self._save_to_bytes(wb)

    def save(self, file_path: str, file_content: bytes) -> None:
        """Save the provided bytes to the specified file path."""
        Path(file_path).write_bytes(file_content)

    def _populate_sheet(
        self,
        ws: Worksheet,
        headers: list[str],
        rows: list[list[CellValue]],
    ) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
        for row_data in rows:
            ws.append([_resolve_cell_value(entry) for entry in row_data])
            row_idx = ws.max_row
            for col_idx, entry in enumerate(row_data, start=1):
                if isinstance(entry, Percent):
                    ws.cell(row=row_idx, column=col_idx).number_format = _PERCENT_NUMBER_FORMAT
        ws.auto_filter.ref = ws.dimensions

    def _save_to_bytes(self, wb: Workbook) -> bytes:
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
