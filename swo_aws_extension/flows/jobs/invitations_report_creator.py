import datetime as dt
import logging
from io import BytesIO

from azure.storage.blob import BlobSasPermissions, BlobServiceClient, generate_blob_sas
from mpt_extension_sdk.mpt_http.base import MPTClient
from mpt_extension_sdk.mpt_http.mpt import _paginated, get_agreements_by_query  # NoQA: PLC2701
from openpyxl import Workbook
from openpyxl.styles import Font

from swo_aws_extension.aws.client import AWSClient
from swo_aws_extension.aws.errors import AWSError
from swo_aws_extension.config import get_config
from swo_aws_extension.parameters import (
    get_channel_handshake_approval_status,
    get_customer_roles_deployed,
    get_mpa_account_id,
    get_responsibility_transfer_id,
    get_support_type,
)
from swo_aws_extension.swo.notifications.teams import Button, TeamsNotificationManager
from swo_aws_extension.swo.rql.query_builder import RQLQuery

logger = logging.getLogger(__name__)


INVITATIONS_REPORT_HEADERS = [
    "PMA Account ID",
    "MPA Account ID",
    "Invitation ID",
    "Invitation Status",
    "Agreement ID",
    "Agreement Name",
    "Product ID",
    "Agreement Status",
    "Support Type",
    "Customer Roles Deployed",
    "Channel Handshake Approval Status",
    "Start Date",
    "End Date",
]


class InvitationsReportCreator:
    """Create invitations report."""

    def __init__(self, mpt_client: MPTClient, product_ids: list[str]):
        self.mpt_client = mpt_client
        self.product_ids = product_ids

    def create_and_notify_teams(self) -> None:
        """Create invitations report, upload to Azure Storage and notify via Teams.

        The Excel is built in memory, uploaded to Azure Blob Storage and a Teams
        notification with a download link (SAS URL valid for a configured period of time) is sent.
        """
        logger.info("Creating invitations report for Teams notification...")
        rows = self._collect_rows()
        excel_bytes = self._build_excel(rows)
        download_url = self._upload_to_azure(excel_bytes)
        self._notify_teams(len(rows), download_url)
        logger.info("Invitations report uploaded and Teams notification sent.")

    def _collect_rows(self) -> list[list[str]]:
        """Fetch agreements and collect report rows."""
        rql_query = RQLQuery(product__id__in=self.product_ids)
        authorizations = self._get_authorizations(self.mpt_client, rql_query, 10)

        rows: list[list[str]] = []
        for authorization in authorizations:
            select = "&select=parameters,authorization.externalIds.operations"
            rql_filter = RQLQuery(authorization__id__eq=authorization["id"]) & RQLQuery(
                status__ne="Draft"
            )
            query_str = f"{rql_filter}{select}"
            agreements = get_agreements_by_query(self.mpt_client, query_str)

            try:
                aws_client = AWSClient(
                    get_config(),
                    authorization["externalIds"]["operations"],
                    get_config().management_role_name,
                )
                aws_invitations = aws_client.get_inbound_responsibility_transfers()
            except AWSError as error:
                logger.warning(
                    "Error getting AWS invitations for authorization %s: %s",
                    authorization["id"],
                    error,
                )
                error_message = (
                    f"Error getting AWS invitations for authorization {authorization['id']}:{error}"
                )
                TeamsNotificationManager().send_error(
                    "Error getting AWS invitations ", error_message
                )
                continue

            for invitation in aws_invitations:
                invitations_row = self._invitation_to_row(invitation, agreements)
                rows.append(invitations_row)

            agreements_without_invitations = self._agreements_without_invitations(
                agreements, aws_invitations
            )
            rows.extend(agreements_without_invitations)

        return rows

    def _invitation_to_row(self, invitation: dict, agreements: dict) -> list[str]:
        start_date = (
            invitation.get("StartTimestamp").strftime("%Y-%m-%d")
            if invitation.get("StartTimestamp")
            else ""
        )
        end_date = (
            invitation.get("EndTimestamp").strftime("%Y-%m-%d")
            if invitation.get("EndTimestamp")
            else ""
        )
        for agreement in agreements:
            invitation_id = get_responsibility_transfer_id(agreement)
            if invitation_id == invitation["Id"]:
                product = agreement.get("product") or {}
                product_id = product.get("id", "")
                support_type = get_support_type(agreement)
                roles = get_customer_roles_deployed(agreement).capitalize()
                handshake_approval_status = get_channel_handshake_approval_status(
                    agreement
                ).capitalize()
                invitation_id = get_responsibility_transfer_id(agreement)
                return [
                    invitation.get("Target", {}).get("ManagementAccountId", ""),
                    invitation.get("Source", {}).get("ManagementAccountId", ""),
                    invitation_id,
                    invitation["Status"],
                    agreement.get("id", ""),
                    agreement.get("name", ""),
                    product_id,
                    agreement.get("status", ""),
                    support_type,
                    roles,
                    handshake_approval_status,
                    start_date,
                    end_date,
                ]
        return [
            invitation.get("Target", {}).get("ManagementAccountId", ""),
            invitation.get("Source", {}).get("ManagementAccountId", ""),
            invitation.get("Id", ""),
            invitation.get("Status", ""),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            start_date,
            end_date,
        ]

    def _agreements_without_invitations(
        self, agreements: list[dict], aws_invitations: list[dict]
    ) -> list[dict]:
        """Return agreements without invitations."""
        agreements_without_invitations = []
        for agreement in agreements:
            invitation_id = get_responsibility_transfer_id(agreement)

            for invitation in aws_invitations:
                if invitation["Id"] == invitation_id:
                    break
            else:
                agreements_without_invitations.append(
                    self._agreement_to_row(agreement, aws_invitations)
                )
        return agreements_without_invitations

    def _agreement_to_row(self, agreement: dict, aws_invitations: list[dict]) -> list[str]:
        """Extract report row from a single agreement."""
        auth = agreement.get("authorization") or {}
        ext_ids = auth.get("externalIds") or {}
        pma_account_id = ext_ids.get("operations", "")
        product = agreement.get("product") or {}
        product_id = product.get("id", "")
        mpa_account_id = get_mpa_account_id(agreement) or ""
        support_type = get_support_type(agreement)
        roles = get_customer_roles_deployed(agreement).capitalize()
        handshake_approval_status = get_channel_handshake_approval_status(agreement).capitalize()
        invitation_id = get_responsibility_transfer_id(agreement)

        return [
            str(pma_account_id),
            str(mpa_account_id),
            invitation_id,
            "",
            agreement.get("id", ""),
            agreement.get("name", ""),
            product_id,
            agreement.get("status", ""),
            support_type,
            roles,
            handshake_approval_status,
            "",
            "",
        ]

    def _build_excel(self, rows: list[list[str]]) -> bytes:
        """Build an Excel workbook in memory and return its raw bytes."""
        wb = Workbook()
        ws = wb.active
        if ws.title == "Sheet":
            ws.title = "Invitations"

        header_font = Font(bold=True)
        for col, header in enumerate(INVITATIONS_REPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        ws.auto_filter.ref = ws.dimensions

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _upload_to_azure(self, excel_data: bytes) -> str:
        """Upload the in-memory Excel to Azure Blob Storage and return a SAS download URL.

        The blob is stored under configured container.

        Args:
            excel_data: Raw bytes of the Excel workbook.

        Returns:
            A publicly accessible SAS URL valid for the number of days specified in config.
        """
        config = get_config()
        connection_string = config.azure_storage_connection_string
        container_name = config.azure_storage_container
        today = dt.datetime.now(dt.UTC).strftime("%Y-%m-%d")
        blob_name = f"{today}.xlsx"
        sas_expiry_days = config.azure_storage_sas_expiry_days

        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        blob_client = blob_service_client.get_blob_client(
            container=container_name,
            blob=blob_name,
        )
        blob_client.upload_blob(excel_data, overwrite=True)
        logger.info("Report uploaded to Azure Blob: %s/%s", container_name, blob_name)

        account_name = blob_service_client.account_name
        account_key = blob_service_client.credential.account_key
        sas_token = generate_blob_sas(
            account_name=account_name,
            container_name=container_name,
            blob_name=blob_name,
            account_key=account_key,
            permission=BlobSasPermissions(read=True),
            expiry=dt.datetime.now(dt.UTC) + dt.timedelta(days=sas_expiry_days),
        )
        return f"{blob_client.url}?{sas_token}"

    def _notify_teams(self, row_count: int, download_url: str) -> None:
        """Send a Teams notification with a download link for the report.

        Args:
            row_count: Number of agreement rows in the report.
            download_url: SAS URL for downloading the report from Azure Storage.
        """
        title = "AWS Billing Transfer Invitations Report"
        text = f"The report has been generated with **{row_count}** agreement(s)."
        button = Button(label="Download report", url=download_url)
        TeamsNotificationManager().send_success(title, text, button=button)

    def _get_authorizations(self, mpt_client, rql_query, limit=10):
        """
        Retrieve authorizations based on the provided RQL query.

        Args:
            mpt_client (MPTClient): MPT API client instance.
            rql_query (RQLQuery): Query to filter authorizations.
            limit (int): Maximum number of authorizations to retrieve.

        Returns:
            list or None: List of authorizations or None if request fails.
        """
        url = (
            f"/catalog/authorizations?{rql_query}&select=externalIds,product"
            if rql_query
            else "/catalog/authorizations?select=externalIds,product"
        )
        return _paginated(mpt_client, url, limit=limit)
